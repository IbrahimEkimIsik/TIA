#!/usr/bin/env python3
"""
Rosti Threat Intelligence CLI
Natural language interface to rosti.dev threat intelligence reports and IOCs.

Usage:
    python rosti.py "find reports about LummaC2"
    python rosti.py "get all IP IOCs from last week"
    python rosti.py "search for malicious domains"
    python rosti.py  (interactive mode if no argument given)
"""

import os
import sys
import json
import requests
import anthropic
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

load_dotenv()

ROSTI_API_KEY = os.environ.get("ROSTI_API_KEY", "")
ROSTI_BASE = "https://rosti.dev/api/v1"
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

rosti_headers = {"X-Api-Key": ROSTI_API_KEY}


# ---------------------------------------------------------------------------
# Rosti API helpers
# ---------------------------------------------------------------------------

def rosti_get(path: str, params: dict = None) -> dict | list:
    url = f"{ROSTI_BASE}{path}"
    resp = requests.get(url, headers=rosti_headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------

def search_reports(q: str) -> str:
    """Search reports by title, authors, or source."""
    data = rosti_get("/search/reports", {"q": q})
    if not data:
        return f"No reports found for '{q}'."
    lines = [f"Found {len(data)} report(s) matching '{q}':\n"]
    for r in data[:20]:
        lines.append(
            f"  [{r['id']}] {r['date']} — {r['title']}\n"
            f"    Source: {r['source']} | Tags: {', '.join(r.get('tags', []))}\n"
            f"    IOCs: {r.get('count', {}).get('iocs', 0)} | YARA: {r.get('count', {}).get('yararules', 0)}\n"
            f"    URL: {r.get('url', '')}"
        )
    if len(data) > 20:
        lines.append(f"\n  ... and {len(data) - 20} more.")
    return "\n".join(lines)


def search_iocs(q: str, pattern: bool = False) -> str:
    """Search IOCs by value (exact or substring match)."""
    data = rosti_get("/search/iocs", {"q": q, "pattern": str(pattern).lower()})
    if not data:
        return f"No IOCs found matching '{q}'."
    lines = [f"Found {len(data)} IOC(s) matching '{q}':\n"]
    for ioc in data[:50]:
        comment = f" — {ioc['comment']}" if ioc.get("comment") else ""
        risk = f" [risk:{ioc['risk']}]" if ioc.get("risk") else ""
        lines.append(
            f"  [{ioc['ioctype']}] {ioc['value']}{comment}{risk}\n"
            f"    Category: {ioc.get('category', 'n/a')} | Report: {ioc.get('report', 'n/a')} | Date: {ioc.get('date', 'n/a')}"
        )
    if len(data) > 50:
        lines.append(f"\n  ... and {len(data) - 50} more.")
    return "\n".join(lines)


def get_reports(fromdate: str = None, source: str = None, limit: int = 20, offset: int = 0) -> str:
    """List reports, optionally filtered by date and/or source."""
    params = {"limit": limit, "offset": offset}
    if fromdate:
        params["fromdate"] = fromdate
    if source:
        params["source"] = source
    data = rosti_get("/reports", params)
    if not data:
        return "No reports found."
    lines = [f"Showing {len(data)} report(s):\n"]
    for r in data:
        lines.append(
            f"  [{r['id']}] {r['date']} — {r['title']}\n"
            f"    Source: {r['source']} | Tags: {', '.join(r.get('tags', []))}\n"
            f"    IOCs: {r.get('count', {}).get('iocs', 0)} | YARA: {r.get('count', {}).get('yararules', 0)}"
        )
    return "\n".join(lines)


def get_report(report_id: str) -> str:
    """Get a specific report by ID including its IOCs and YARA rules."""
    data = rosti_get(f"/reports/{report_id}")
    report = data  # API returns report fields at top level
    iocs = data.get("iocs", [])
    yara = data.get("yararules", [])

    lines = [
        f"Report: {report.get('title', 'Unknown')}",
        f"  ID: {report.get('id')} | Date: {report.get('date')} | Source: {report.get('source')}",
        f"  Authors: {', '.join(report.get('authors', []))}",
        f"  Tags: {', '.join(report.get('tags', []))}",
        f"  URL: {report.get('url', '')}",
        f"  IOCs: {len(iocs)} | YARA rules: {len(yara)}",
        "",
    ]

    if iocs:
        lines.append(f"IOCs ({len(iocs)} total):")
        by_type: dict = {}
        for ioc in iocs:
            by_type.setdefault(ioc["ioctype"], []).append(ioc)
        for ioc_type, items in sorted(by_type.items()):
            lines.append(f"\n  [{ioc_type.upper()}] ({len(items)} items)")
            for ioc in items[:30]:
                comment = f" — {ioc['comment']}" if ioc.get("comment") else ""
                lines.append(f"    {ioc['value']}{comment}")
            if len(items) > 30:
                lines.append(f"    ... and {len(items) - 30} more")

    if yara:
        lines.append(f"\nYARA Rules ({len(yara)} total):")
        for rule in yara[:5]:
            lines.append(f"  - {rule['name']}")
        if len(yara) > 5:
            lines.append(f"  ... and {len(yara) - 5} more")

    return "\n".join(lines)


def get_iocs(ioctype: str = None, category: str = None, limit: int = 50, offset: int = 0) -> str:
    """List IOCs, optionally filtered by type and/or category."""
    params = {"limit": limit, "offset": offset}
    if ioctype:
        params["ioctype"] = ioctype
    if category:
        params["category"] = category
    data = rosti_get("/iocs", params)
    if not data:
        return "No IOCs found."
    lines = [f"Showing {len(data)} IOC(s):\n"]
    for ioc in data:
        comment = f" — {ioc['comment']}" if ioc.get("comment") else ""
        lines.append(
            f"  [{ioc['ioctype']}] {ioc['value']}{comment}\n"
            f"    Category: {ioc.get('category', 'n/a')} | Report: {ioc.get('report', 'n/a')} | Date: {ioc.get('date', 'n/a')}"
        )
    return "\n".join(lines)


def export_iocs_to_excel(report_ids: list, filename: str) -> str:
    """Fetch IOCs from multiple reports and export to an Excel file, one sheet per report."""
    if not filename.endswith(".xlsx"):
        filename = filename.replace(".xls", "") + ".xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")

    exported = []
    skipped = []

    for report_id in report_ids:
        try:
            data = rosti_get(f"/reports/{report_id}")
        except Exception as e:
            skipped.append(f"{report_id} (error: {e})")
            continue

        report = data  # API returns report fields at top level
        iocs = data.get("iocs", [])

        if not iocs:
            skipped.append(f"{report_id} (no IOCs)")
            continue

        # Sheet name: max 31 chars, no special chars
        title = report.get("title", report_id)
        for ch in r'\/*?:[]':
            title = title.replace(ch, "")
        title = title.strip()
        if len(title) > 31:
            # Truncate at last word boundary within 31 chars
            truncated = title[:31].rsplit(" ", 1)[0]
            title = truncated if truncated else title[:31]
        sheet_name = title if title else report_id[:31]

        # Avoid duplicate sheet names
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:29] + str(len(existing))

        ws = wb.create_sheet(title=sheet_name)

        # Metadata rows
        ws.append(["Report ID", report.get("id", "")])
        ws.append(["Title", report.get("title", "")])
        ws.append(["Date", report.get("date", "")])
        ws.append(["Source", report.get("source", "")])
        ws.append(["URL", report.get("url", "")])
        ws.append([])

        # IOC table headers
        headers = ["Type", "Value", "Comment", "Risk", "Category", "Date"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # IOC rows
        for ioc in iocs:
            ws.append([
                ioc.get("ioctype", ""),
                ioc.get("value", ""),
                ioc.get("comment", ""),
                ioc.get("risk", ""),
                ioc.get("category", ""),
                ioc.get("date", ""),
            ])

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        exported.append(f"{report_id} — {report.get('title', '')} ({len(iocs)} IOCs)")

    if not exported:
        return "No IOCs found for any of the provided report IDs. Nothing exported."

    wb.save(filename)
    lines = [f"Exported {len(exported)} sheet(s) to '{filename}':\n"]
    for e in exported:
        lines.append(f"  + {e}")
    if skipped:
        lines.append(f"\nSkipped ({len(skipped)}):")
        for s in skipped:
            lines.append(f"  - {s}")
    return "\n".join(lines)


def export_iocs_to_csv(report_ids: list, filename: str) -> str:
    """Fetch IOCs from multiple reports and export to a CSV file."""
    import csv
    if not filename.endswith(".csv"):
        filename += ".csv"
    rows = []
    skipped = []
    for report_id in report_ids:
        try:
            data = rosti_get(f"/reports/{report_id}")
        except Exception as e:
            skipped.append(f"{report_id} (error: {e})")
            continue
        iocs = data.get("iocs", [])
        if not iocs:
            skipped.append(f"{report_id} (no IOCs)")
            continue
        for ioc in iocs:
            rows.append({
                "report_id": data.get("id", ""),
                "report_title": data.get("title", ""),
                "report_date": data.get("date", ""),
                "report_source": data.get("source", ""),
                "report_url": data.get("url", ""),
                "ioctype": ioc.get("ioctype", ""),
                "value": ioc.get("value", ""),
                "comment": ioc.get("comment", ""),
                "risk": ioc.get("risk", ""),
                "category": ioc.get("category", ""),
                "date": ioc.get("date", ""),
            })
    if not rows:
        return "No IOCs found. Nothing exported."
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    result = f"Exported {len(rows)} IOCs from {len(report_ids) - len(skipped)} report(s) to '{filename}'."
    if skipped:
        result += f"\nSkipped: {', '.join(skipped)}"
    return result


def export_iocs_to_json(report_ids: list, filename: str) -> str:
    """Fetch IOCs from multiple reports and export to a JSON file."""
    if not filename.endswith(".json"):
        filename += ".json"
    output = []
    skipped = []
    for report_id in report_ids:
        try:
            data = rosti_get(f"/reports/{report_id}")
        except Exception as e:
            skipped.append(f"{report_id} (error: {e})")
            continue
        iocs = data.get("iocs", [])
        if not iocs:
            skipped.append(f"{report_id} (no IOCs)")
            continue
        output.append({
            "report_id": data.get("id", ""),
            "report_title": data.get("title", ""),
            "report_date": data.get("date", ""),
            "report_source": data.get("source", ""),
            "report_url": data.get("url", ""),
            "iocs": iocs,
        })
    if not output:
        return "No IOCs found. Nothing exported."
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    total_iocs = sum(len(r["iocs"]) for r in output)
    result = f"Exported {total_iocs} IOCs from {len(output)} report(s) to '{filename}'."
    if skipped:
        result += f"\nSkipped: {', '.join(skipped)}"
    return result


def get_sources() -> str:
    """List all available threat intel sources."""
    data = rosti_get("/sources")
    lines = [f"Available sources ({len(data)}):\n"]
    for s in data:
        lines.append(f"  {s['id']:30s} {s['name']} — {s.get('url', '')}")
    return "\n".join(lines)


def get_ioc_types() -> str:
    """List all available IOC types."""
    data = rosti_get("/ioctypes")
    return f"IOC types: {', '.join(data)}"


def get_categories() -> str:
    """List all available IOC categories."""
    data = rosti_get("/categories")
    return f"IOC categories: {', '.join(data)}"


# ---------------------------------------------------------------------------
# Tool definitions for Claude
# ---------------------------------------------------------------------------

TOOLS = [
    {
        "name": "search_reports",
        "description": "Search threat intelligence reports by title, authors, or source name. Use this to find reports about specific malware families, threat actors, or topics.",
        "input_schema": {
            "type": "object",
            "properties": {
                "q": {"type": "string", "description": "Search query (e.g. 'LummaC2', 'ransomware', 'APT29')"}
            },
            "required": ["q"]
        }
    },
    {
        "name": "search_iocs",
        "description": "Search for specific IOC values (IPs, domains, hashes, URLs). Use exact=true for exact match, pattern=true to search substrings.",
        "input_schema": {
            "type": "object",
            "properties": {
                "q": {"type": "string", "description": "IOC value to search for (e.g. 'example.com', '192.168.1.1')"},
                "pattern": {"type": "boolean", "description": "If true, returns IOCs containing the query as a substring. Default: false (exact match)."}
            },
            "required": ["q"]
        }
    },
    {
        "name": "get_reports",
        "description": "List threat intelligence reports, optionally filtered by date and/or source. Use to get recent reports or reports from a specific source.",
        "input_schema": {
            "type": "object",
            "properties": {
                "fromdate": {"type": "string", "description": "Only return reports from or after this date (YYYY-MM-DD format). Example: '2024-01-01'"},
                "source": {"type": "string", "description": "Filter by source ID (e.g. 'malwarebytes', 'prodaft'). Use get_sources to see all source IDs."},
                "limit": {"type": "integer", "description": "Number of reports to return (max 1000, default 20)"},
                "offset": {"type": "integer", "description": "Number of reports to skip (for pagination, default 0)"}
            },
            "required": []
        }
    },
    {
        "name": "get_report",
        "description": "Get a specific report by its ID including all IOCs and YARA rules. Use after search_reports to get full details.",
        "input_schema": {
            "type": "object",
            "properties": {
                "report_id": {"type": "string", "description": "The 8-character report ID (e.g. 'dq13KaQv')"}
            },
            "required": ["report_id"]
        }
    },
    {
        "name": "get_iocs",
        "description": "List IOCs filtered by type (domain, ip, url, hash, etc.) and/or category. Use to bulk-retrieve IOCs of a specific type.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ioctype": {"type": "string", "description": "Filter by IOC type (e.g. 'domain', 'ip', 'url', 'md5', 'sha256'). Use get_ioc_types to see all."},
                "category": {"type": "string", "description": "Filter by category (e.g. 'network_activity', 'payload_delivery'). Use get_categories to see all."},
                "limit": {"type": "integer", "description": "Number of IOCs to return (max 1000, default 50)"},
                "offset": {"type": "integer", "description": "Number of IOCs to skip (for pagination, default 0)"}
            },
            "required": []
        }
    },
    {
        "name": "get_sources",
        "description": "List all available threat intelligence sources with their IDs and URLs.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": []
        }
    },
    {
        "name": "get_ioc_types",
        "description": "List all available IOC types (domain, ip, url, hash, etc.).",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": []
        }
    },
    {
        "name": "get_categories",
        "description": "List all available IOC categories.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": []
        }
    },
    {
        "name": "export_iocs_to_excel",
        "description": "Fetch IOCs from multiple reports and export them to an Excel (.xlsx) file, with one sheet per report. Use this when the user asks to export or save IOCs to Excel or .xls format.",
        "input_schema": {
            "type": "object",
            "properties": {
                "report_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of report IDs to fetch IOCs from (e.g. ['dq13KaQv', 'abc12345'])"
                },
                "filename": {
                    "type": "string",
                    "description": "Output filename (e.g. 'dprk_iocs_2026.xlsx')"
                }
            },
            "required": ["report_ids", "filename"]
        }
    },
    {
        "name": "export_iocs_to_csv",
        "description": "Fetch IOCs from multiple reports and export to a CSV file. Use ONLY when the user explicitly asks to export or save to CSV.",
        "input_schema": {
            "type": "object",
            "properties": {
                "report_ids": {"type": "array", "items": {"type": "string"}, "description": "List of report IDs"},
                "filename": {"type": "string", "description": "Output filename (e.g. 'iocs.csv')"}
            },
            "required": ["report_ids", "filename"]
        }
    },
    {
        "name": "export_iocs_to_json",
        "description": "Fetch IOCs from multiple reports and export to a JSON file. Use ONLY when the user explicitly asks to export or save to JSON.",
        "input_schema": {
            "type": "object",
            "properties": {
                "report_ids": {"type": "array", "items": {"type": "string"}, "description": "List of report IDs"},
                "filename": {"type": "string", "description": "Output filename (e.g. 'iocs.json')"}
            },
            "required": ["report_ids", "filename"]
        }
    },
]

TOOL_MAP = {
    "search_reports": lambda inp: search_reports(**inp),
    "search_iocs": lambda inp: search_iocs(**inp),
    "get_reports": lambda inp: get_reports(**inp),
    "get_report": lambda inp: get_report(**inp),
    "get_iocs": lambda inp: get_iocs(**inp),
    "get_sources": lambda inp: get_sources(),
    "get_ioc_types": lambda inp: get_ioc_types(),
    "get_categories": lambda inp: get_categories(),
    "export_iocs_to_excel": lambda inp: export_iocs_to_excel(**inp),
    "export_iocs_to_csv": lambda inp: export_iocs_to_csv(**inp),
    "export_iocs_to_json": lambda inp: export_iocs_to_json(**inp),
}

SYSTEM_PROMPT = """You are a threat intelligence analyst assistant with access to the Rosti threat intelligence platform.
Rosti aggregates open-source threat intelligence reports and IOCs (Indicators of Compromise) from security vendors.

When the user asks about threat intelligence, malware, IOCs, or security reports:
1. Use the available tools to query the Rosti API
2. ALWAYS present reports in a markdown table with exactly 3 columns: Date | Report Name | URL. The URL column must always contain the full raw URL (e.g. https://...). Never truncate or hide the URL. Example:
   | Date | Report Name | URL |
   |------|-------------|-----|
   | 2026-02-24 | North Korean Lazarus Group Now Working With Medusa Ransomware | https://www.security.com/threat-intelligence/lazarus-medusa-ransomware |
3. Highlight the most relevant findings after the table
4. If a search returns report IDs, use get_report to fetch details when the user wants IOCs
5. ONLY export to a file (Excel, CSV, or JSON) when the user explicitly asks for it. If the user does not mention exporting or saving to a file, never call export_iocs_to_excel — just show results in the terminal.

Common IOC types: domain, ip, url, md5, sha256, sha1, email, filename, registry
Common categories: network_activity, payload_delivery, artifacts_dropped, external_analysis

Be concise but complete. Format IOC lists clearly."""


def run_query(prompt: str) -> None:
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    messages = [{"role": "user", "content": prompt}]

    print(f"\nQuerying Rosti for: {prompt}\n{'=' * 60}")

    while True:
        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=4096,
            system=SYSTEM_PROMPT,
            tools=TOOLS,
            messages=messages,
        )

        # Print any text content
        for block in response.content:
            if block.type == "text":
                print(block.text)

        if response.stop_reason == "end_turn":
            break

        if response.stop_reason != "tool_use":
            break

        # Execute tool calls
        messages.append({"role": "assistant", "content": response.content})
        tool_results = []

        for block in response.content:
            if block.type != "tool_use":
                continue

            tool_name = block.name
            tool_input = block.input
            print(f"\n[Calling: {tool_name}({json.dumps(tool_input, ensure_ascii=False)})]")

            try:
                result = TOOL_MAP[tool_name](tool_input)
            except requests.HTTPError as e:
                result = f"API error: {e.response.status_code} — {e.response.text}"
            except Exception as e:
                result = f"Error: {e}"

            tool_results.append({
                "type": "tool_result",
                "tool_use_id": block.id,
                "content": result,
            })

        messages.append({"role": "user", "content": tool_results})


def main():
    if not ROSTI_API_KEY:
        print("Error: ROSTI_API_KEY not set in .env")
        sys.exit(1)

    if not ANTHROPIC_API_KEY or ANTHROPIC_API_KEY == "your_anthropic_api_key_here":
        print("Error: ANTHROPIC_API_KEY not set in .env")
        sys.exit(1)

    if len(sys.argv) > 1:
        prompt = " ".join(sys.argv[1:])
        run_query(prompt)
    else:
        print("Rosti Threat Intelligence CLI")
        print("Type your query or 'quit' to exit.\n")
        while True:
            try:
                prompt = input(">> ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nBye.")
                break
            if not prompt:
                continue
            if prompt.lower() in ("quit", "exit", "q"):
                print("Bye.")
                break
            run_query(prompt)


if __name__ == "__main__":
    main()
