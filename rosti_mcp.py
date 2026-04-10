#!/usr/bin/env python3
"""
Rosti Threat Intelligence MCP Server
Exposes rosti.dev API as native tools for Claude Code.
"""

import os
import csv
import json
import io
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

ROSTI_API_KEY = os.environ.get("ROSTI_API_KEY", "")
ROSTI_BASE = "https://rosti.dev/api/v1"
rosti_headers = {"X-Api-Key": ROSTI_API_KEY}

mcp = FastMCP("rosti")


# ---------------------------------------------------------------------------
# API helper
# ---------------------------------------------------------------------------

def rosti_get(path: str, params: dict = None) -> dict | list:
    url = f"{ROSTI_BASE}{path}"
    resp = requests.get(url, headers=rosti_headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------

@mcp.tool()
def search_reports(q: str) -> str:
    """Search threat intelligence reports by title, authors, or source name.
    Use this to find reports about specific malware families, threat actors, or topics.

    Args:
        q: Search query (e.g. 'LummaC2', 'Lazarus', 'DPRK', 'Scattered Spider')
    """
    data = rosti_get("/search/reports", {"q": q})
    if not data:
        return f"No reports found for '{q}'."
    lines = [f"Found {len(data)} report(s) matching '{q}':\n"]
    for r in data:
        lines.append(
            f"  ID: {r['id']} | Date: {r['date']} | IOCs: {r.get('count', {}).get('iocs', 0)}\n"
            f"  Title: {r['title']}\n"
            f"  URL: {r.get('url', '')}\n"
        )
    return "\n".join(lines)


@mcp.tool()
def get_reports(fromdate: str = None, source: str = None, limit: int = 20, offset: int = 0) -> str:
    """List threat intelligence reports, optionally filtered by date and/or source.

    Args:
        fromdate: Only return reports from or after this date (YYYY-MM-DD). Example: '2026-01-01'
        source: Filter by source ID (e.g. 'malwarebytes', 'prodaft')
        limit: Number of reports to return (max 1000, default 20)
        offset: Number of reports to skip for pagination (default 0)
    """
    params = {"limit": limit, "offset": offset}
    if fromdate:
        params["fromdate"] = fromdate
    if source:
        params["source"] = source
    data = rosti_get("/reports", params)
    if not data:
        return "No reports found."
    lines = [f"Showing {len(data)} report(s):\n"]
    for r in data:
        lines.append(
            f"  ID: {r['id']} | Date: {r['date']} | IOCs: {r.get('count', {}).get('iocs', 0)}\n"
            f"  Title: {r['title']}\n"
            f"  URL: {r.get('url', '')}\n"
        )
    return "\n".join(lines)


@mcp.tool()
def get_report(report_id: str) -> str:
    """Get a specific report by ID including all its IOCs and YARA rules.

    Args:
        report_id: The 8-character report ID (e.g. 'dq13KaQv')
    """
    data = rosti_get(f"/reports/{report_id}")
    iocs = data.get("iocs", [])
    yara = data.get("yararules", [])

    lines = [
        f"Title: {data.get('title', 'Unknown')}",
        f"ID: {data.get('id')} | Date: {data.get('date')} | Source: {data.get('source')}",
        f"Authors: {', '.join(data.get('authors', []))}",
        f"Tags: {', '.join(data.get('tags', []))}",
        f"URL: {data.get('url', '')}",
        f"IOCs: {len(iocs)} | YARA rules: {len(yara)}",
        "",
    ]

    if iocs:
        lines.append(f"IOCs ({len(iocs)} total):")
        by_type: dict = {}
        for ioc in iocs:
            by_type.setdefault(ioc["ioctype"], []).append(ioc)
        for ioc_type, items in sorted(by_type.items()):
            lines.append(f"\n  [{ioc_type.upper()}] ({len(items)} items)")
            for ioc in items[:30]:
                comment = f" — {ioc['comment']}" if ioc.get("comment") else ""
                lines.append(f"    {ioc['value']}{comment}")
            if len(items) > 30:
                lines.append(f"    ... and {len(items) - 30} more")

    if yara:
        lines.append(f"\nYARA Rules ({len(yara)} total):")
        for rule in yara[:5]:
            lines.append(f"  - {rule['name']}")
        if len(yara) > 5:
            lines.append(f"  ... and {len(yara) - 5} more")

    return "\n".join(lines)


@mcp.tool()
def search_iocs(q: str, pattern: bool = False) -> str:
    """Search for specific IOC values (IPs, domains, hashes, URLs).

    Args:
        q: IOC value to search for (e.g. 'example.com', '192.168.1.1')
        pattern: If true, returns IOCs containing the query as a substring. Default: false (exact match)
    """
    data = rosti_get("/search/iocs", {"q": q, "pattern": str(pattern).lower()})
    if not data:
        return f"No IOCs found matching '{q}'."
    lines = [f"Found {len(data)} IOC(s) matching '{q}':\n"]
    for ioc in data[:50]:
        comment = f" — {ioc['comment']}" if ioc.get("comment") else ""
        risk = f" [risk:{ioc['risk']}]" if ioc.get("risk") else ""
        lines.append(
            f"  [{ioc['ioctype']}] {ioc['value']}{comment}{risk}\n"
            f"    Category: {ioc.get('category', 'n/a')} | Report: {ioc.get('report', 'n/a')} | Date: {ioc.get('date', 'n/a')}"
        )
    if len(data) > 50:
        lines.append(f"\n  ... and {len(data) - 50} more.")
    return "\n".join(lines)


@mcp.tool()
def get_iocs(ioctype: str = None, category: str = None, limit: int = 50, offset: int = 0) -> str:
    """List IOCs filtered by type and/or category.

    Args:
        ioctype: Filter by IOC type (e.g. 'domain', 'ip', 'url', 'md5', 'sha256')
        category: Filter by category (e.g. 'network_activity', 'payload_delivery')
        limit: Number of IOCs to return (max 1000, default 50)
        offset: Number of IOCs to skip for pagination (default 0)
    """
    params = {"limit": limit, "offset": offset}
    if ioctype:
        params["ioctype"] = ioctype
    if category:
        params["category"] = category
    data = rosti_get("/iocs", params)
    if not data:
        return "No IOCs found."
    lines = [f"Showing {len(data)} IOC(s):\n"]
    for ioc in data:
        comment = f" — {ioc['comment']}" if ioc.get("comment") else ""
        lines.append(
            f"  [{ioc['ioctype']}] {ioc['value']}{comment}\n"
            f"    Category: {ioc.get('category', 'n/a')} | Report: {ioc.get('report', 'n/a')} | Date: {ioc.get('date', 'n/a')}"
        )
    return "\n".join(lines)


@mcp.tool()
def get_sources() -> str:
    """List all available threat intelligence sources with their IDs and URLs."""
    data = rosti_get("/sources")
    lines = [f"Available sources ({len(data)}):\n"]
    for s in data:
        lines.append(f"  {s['id']:30s} {s['name']} — {s.get('url', '')}")
    return "\n".join(lines)


@mcp.tool()
def get_ioc_types() -> str:
    """List all available IOC types (domain, ip, url, hash, etc.)."""
    data = rosti_get("/ioctypes")
    return f"IOC types: {', '.join(data)}"


@mcp.tool()
def get_categories() -> str:
    """List all available IOC categories."""
    data = rosti_get("/categories")
    return f"IOC categories: {', '.join(data)}"


@mcp.tool()
def export_iocs_to_excel(report_ids: list[str], filename: str) -> str:
    """Fetch IOCs from multiple reports and export to an Excel (.xlsx) file, one sheet per report.
    Use ONLY when the user explicitly asks to export to Excel or .xlsx/.xls format.

    Args:
        report_ids: List of report IDs to fetch IOCs from (e.g. ['dq13KaQv', 'abc12345'])
        filename: Output filename (e.g. 'lazarus_iocs_2026.xlsx')
    """
    if not filename.endswith(".xlsx"):
        filename = filename.replace(".xls", "") + ".xlsx"

    output_path = os.path.join(os.path.dirname(__file__), filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")

    exported = []
    skipped = []

    for report_id in report_ids:
        try:
            data = rosti_get(f"/reports/{report_id}")
        except Exception as e:
            skipped.append(f"{report_id} (error: {e})")
            continue

        iocs = data.get("iocs", [])
        if not iocs:
            skipped.append(f"{report_id} (no IOCs)")
            continue

        title = data.get("title", report_id)
        for ch in r'\/*?:[]':
            title = title.replace(ch, "")
        title = title.strip()
        if len(title) > 31:
            truncated = title[:31].rsplit(" ", 1)[0]
            title = truncated if truncated else title[:31]
        sheet_name = title if title else report_id[:31]

        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:29] + str(len(existing))

        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Report ID", data.get("id", "")])
        ws.append(["Title", data.get("title", "")])
        ws.append(["Date", data.get("date", "")])
        ws.append(["Source", data.get("source", "")])
        ws.append(["URL", data.get("url", "")])
        ws.append([])

        headers = ["Type", "Value", "Comment", "Risk", "Category", "Date"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for ioc in iocs:
            ws.append([
                ioc.get("ioctype", ""),
                ioc.get("value", ""),
                ioc.get("comment", ""),
                ioc.get("risk", ""),
                ioc.get("category", ""),
                ioc.get("date", ""),
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        exported.append(f"{report_id} — {data.get('title', '')} ({len(iocs)} IOCs)")

    if not exported:
        return "No IOCs found for any of the provided report IDs. Nothing exported."

    wb.save(output_path)
    lines = [f"Exported {len(exported)} sheet(s) to '{output_path}':\n"]
    for e in exported:
        lines.append(f"  + {e}")
    if skipped:
        lines.append(f"\nSkipped ({len(skipped)}):")
        for s in skipped:
            lines.append(f"  - {s}")
    return "\n".join(lines)


@mcp.tool()
def export_iocs_to_csv(report_ids: list[str], filename: str) -> str:
    """Fetch IOCs from multiple reports and export to a CSV file.
    Use ONLY when the user explicitly asks to export to CSV.

    Args:
        report_ids: List of report IDs to fetch IOCs from
        filename: Output filename (e.g. 'iocs.csv')
    """
    if not filename.endswith(".csv"):
        filename += ".csv"
    output_path = os.path.join(os.path.dirname(__file__), filename)
    rows = []
    skipped = []
    for report_id in report_ids:
        try:
            data = rosti_get(f"/reports/{report_id}")
        except Exception as e:
            skipped.append(f"{report_id} (error: {e})")
            continue
        iocs = data.get("iocs", [])
        if not iocs:
            skipped.append(f"{report_id} (no IOCs)")
            continue
        for ioc in iocs:
            rows.append({
                "report_id": data.get("id", ""),
                "report_title": data.get("title", ""),
                "report_date": data.get("date", ""),
                "report_source": data.get("source", ""),
                "report_url": data.get("url", ""),
                "ioctype": ioc.get("ioctype", ""),
                "value": ioc.get("value", ""),
                "comment": ioc.get("comment", ""),
                "risk": ioc.get("risk", ""),
                "category": ioc.get("category", ""),
                "date": ioc.get("date", ""),
            })
    if not rows:
        return "No IOCs found. Nothing exported."
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    result = f"Exported {len(rows)} IOCs from {len(report_ids) - len(skipped)} report(s) to '{output_path}'."
    if skipped:
        result += f"\nSkipped: {', '.join(skipped)}"
    return result


@mcp.tool()
def export_iocs_to_json(report_ids: list[str], filename: str) -> str:
    """Fetch IOCs from multiple reports and export to a JSON file.
    Use ONLY when the user explicitly asks to export to JSON.

    Args:
        report_ids: List of report IDs to fetch IOCs from
        filename: Output filename (e.g. 'iocs.json')
    """
    if not filename.endswith(".json"):
        filename += ".json"
    output_path = os.path.join(os.path.dirname(__file__), filename)
    output = []
    skipped = []
    for report_id in report_ids:
        try:
            data = rosti_get(f"/reports/{report_id}")
        except Exception as e:
            skipped.append(f"{report_id} (error: {e})")
            continue
        iocs = data.get("iocs", [])
        if not iocs:
            skipped.append(f"{report_id} (no IOCs)")
            continue
        output.append({
            "report_id": data.get("id", ""),
            "report_title": data.get("title", ""),
            "report_date": data.get("date", ""),
            "report_source": data.get("source", ""),
            "report_url": data.get("url", ""),
            "iocs": iocs,
        })
    if not output:
        return "No IOCs found. Nothing exported."
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    total_iocs = sum(len(r["iocs"]) for r in output)
    result = f"Exported {total_iocs} IOCs from {len(output)} report(s) to '{output_path}'."
    if skipped:
        result += f"\nSkipped: {', '.join(skipped)}"
    return result


if __name__ == "__main__":
    mcp.run(transport="stdio")
